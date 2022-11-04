import openpyxl
# from ConfigData.ExcelData import *
def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)
def get_columns_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)
def read_data(file, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.cell(row=rownum, column=columnum).value)